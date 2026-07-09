"""
JPX需給データ プローブスクリプト (一時ファイル・Phase1調査専用)

サンドボックス(Claude Code)からはJPXにproxy403で到達不可なため、実際のURL形式・
ファイル形式・パース可否をGitHub Actionsランナー上で確認するための使い捨てスクリプト。
検証が終わったらリポジトリから削除する(本番実装は jp_supply_demand.py に別途書く)。

確認項目:
  1. 空売り残高報告(0.5%以上, 日次) 一覧ページ → 当日分.xlsリンクの実URL・拡張子・マジックバイト
  2. 銘柄別信用取引週末残高 → PDFかどうかの再確認
  3. 空売り比率(市場全体日次) → PDFかどうかの再確認
  4. 日証金 貸借取引情報(taisyaku.jp) → 生HTMLに表データが載っているか(SPAなら載らない)
"""

import re
import socket
import sys
import urllib.error
import urllib.request

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

socket.setdefaulttimeout(35)

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")


def fetch(url, timeout=20):
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.status, dict(r.getheaders()), r.read()


def probe_short_positions():
    print("\n=== [1] 空売り残高報告(0.5%以上・日次) 一覧ページ ===")
    idx_url = "https://www.jpx.co.jp/markets/public/short-selling/index.html"
    try:
        status, headers, body = fetch(idx_url)
        html = body.decode("utf-8", errors="replace")
        print(f"index status={status} bytes={len(body)}")
        links = re.findall(r'href="([^"]*Short_Positions\.xls[^"]*)"', html, re.IGNORECASE)
        links += re.findall(r'href="([^"]*short-selling[^"]*-att/[^"]*\.xlsx?)"', html, re.IGNORECASE)
        links = sorted(set(links))
        print(f"見つかったxlsリンク数: {len(links)}")
        for l in links[:5]:
            print("  ", l)
        if not links:
            print("index.htmlの生HTML冒頭500字:")
            print(html[:500])
            return
        target = links[0]
        if target.startswith("/"):
            target = "https://www.jpx.co.jp" + target
        elif not target.startswith("http"):
            target = "https://www.jpx.co.jp/markets/public/short-selling/" + target
        print(f"ダウンロード試行: {target}")
        status2, headers2, body2 = fetch(target)
        print(f"status={status2} bytes={len(body2)} content-type={headers2.get('Content-Type')}")
        magic = body2[:8]
        print(f"先頭8バイト(hex): {magic.hex()}")
        if magic[:4] == b'PK\x03\x04':
            print("判定: ZIP系コンテナ → xlsx(OOXML)形式の可能性が高い → openpyxlで読める可能性")
        elif magic[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            print("判定: OLE2/BIFF形式 → 旧来のxls形式 → xlrd等が必要(openpyxl非対応)")
        else:
            print("判定: 不明な形式")
        # 実際に開けるか試す
        with open("/tmp/short_positions_probe.bin", "wb") as f:
            f.write(body2)
        try:
            import openpyxl
            wb = openpyxl.load_workbook("/tmp/short_positions_probe.bin", read_only=True)
            ws = wb[wb.sheetnames[0]]
            print(f"openpyxlで読込成功: sheet={wb.sheetnames}, 先頭行例:")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 5:
                    break
                print("   ", row)
        except Exception as e:
            print(f"openpyxlでの読込失敗: {type(e).__name__}: {e}")
            try:
                import xlrd
                wb = xlrd.open_workbook("/tmp/short_positions_probe.bin")
                ws = wb.sheet_by_index(0)
                print(f"xlrdで読込成功: sheet0 rows={ws.nrows} cols={ws.ncols}")
                for i in range(min(12, ws.nrows)):
                    print("   row", i, ws.row_values(i))
                # ヘッダー行を探す(「銘柄コード」「Code」等を含む行)
                header_i = None
                for i in range(min(20, ws.nrows)):
                    rv = [str(x) for x in ws.row_values(i)]
                    if any("コード" in x or "Code" in x for x in rv):
                        header_i = i
                        print(f"ヘッダー候補行 {i}: {rv}")
                        break
                # universe銘柄コード(7203トヨタ等)を含む行を探して実データ例を出す
                target_codes = {"7203", "6758", "9984", "8306"}
                found = 0
                for i in range(ws.nrows):
                    rv = ws.row_values(i)
                    rvs = {str(int(x)) if isinstance(x, float) and x == int(x) else str(x) for x in rv}
                    if rvs & target_codes:
                        print(f"データ例 row {i}: {rv}")
                        found += 1
                        if found >= 6:
                            break
                print(f"universe銘柄がヒットした行数(先頭一致のみ確認): {found}")
            except Exception as e2:
                print(f"xlrdでの読込も失敗: {type(e2).__name__}: {e2}")
    except Exception as e:
        print(f"[1] 失敗: {type(e).__name__}: {e}")


def probe_margin_weekly():
    print("\n=== [2] 銘柄別信用取引週末残高(週次) ===")
    for url in [
        "https://www.jpx.co.jp/markets/statistics-equities/margin/05.html",
    ]:
        try:
            status, headers, body = fetch(url)
            html = body.decode("utf-8", errors="replace")
            print(f"{url} status={status} bytes={len(body)}")
            links = re.findall(r'href="([^"]*syumatsu[^"]*)"', html, re.IGNORECASE)
            links = sorted(set(links))
            print(f"syumatsuリンク数: {len(links)}")
            for l in links[:5]:
                print("  ", l)
            if links:
                t = links[0]
                if t.startswith("/"):
                    t = "https://www.jpx.co.jp" + t
                s2, h2, b2 = fetch(t)
                print(f"実体ファイル: status={s2} bytes={len(b2)} content-type={h2.get('Content-Type')} "
                      f"先頭4バイト={b2[:4]!r}")
        except Exception as e:
            print(f"[2] 失敗: {type(e).__name__}: {e}")


def probe_short_ratio():
    print("\n=== [3] 空売り比率(市場全体・日次) ===")
    url = "https://www.jpx.co.jp/markets/statistics-equities/short-selling/index.html"
    try:
        status, headers, body = fetch(url)
        html = body.decode("utf-8", errors="replace")
        print(f"{url} status={status} bytes={len(body)}")
        links = re.findall(r'href="([^"]*short-selling[^"]*-att/[^"]*\.(?:pdf|csv|xlsx?))"', html, re.IGNORECASE)
        links = sorted(set(links))
        print(f"リンク数: {len(links)}")
        for l in links[:8]:
            print("  ", l)
    except Exception as e:
        print(f"[3] 失敗: {type(e).__name__}: {e}")


def probe_taisyaku():
    print("\n=== [4] 日証金 貸借取引情報 (taisyaku.jp・補助候補) ===")
    for url in [
        "https://www.taisyaku.jp/download/",
        "https://www.taisyaku.jp/app/stock/detail/7203-01",
    ]:
        try:
            status, headers, body = fetch(url)
            html = body.decode("utf-8", errors="replace")
            print(f"{url} status={status} bytes={len(body)}")
            has_table_digits = bool(re.search(r'<table[\s\S]{0,2000}?\d{3,}', html))
            looks_spa = ('id="root"' in html or 'id="app"' in html) and len(html) < 20000
            print(f"  table+数字を含むか={has_table_digits} / SPAシェルらしき兆候={looks_spa}")
            print("  本文冒頭300字:", re.sub(r'\s+', ' ', html[:300]))
        except Exception as e:
            print(f"[4] 失敗: {type(e).__name__}: {e}")


if __name__ == "__main__":
    probe_short_positions()
    probe_margin_weekly()
    probe_short_ratio()
    probe_taisyaku()
    print("\n=== プローブ完了 ===")
