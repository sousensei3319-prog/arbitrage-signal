"""
850銘柄化 Phase1調査用プローブ (一時ファイル・削除予定)

TOPIX500 / 日経225 の構成銘柄一覧を機械取得できるかを、実URL・実ファイル形式で
確認するための使い捨てスクリプト。サンドボックス(Claude Code)からはJPX/Nikkei
公式サイトにproxy403で到達不可なため、GitHub Actionsランナー上での実行結果でのみ
判断できる (jp_supply_demand.py Phase1調査と同じ制約・同じ手法)。

確認項目:
  1. JPX「東証上場銘柄一覧」(data_j.xls 相当) — 全上場銘柄+規模区分(TOPIX Core30/
     Large70/Mid400等)+33業種区分を持つとされる配布ファイル。TOPIX500 = Core30+
     Large70+Mid400の合算という仮説を実データで検証する。
  2. 日経225 公式構成銘柄ウエイトCSV (indexes.nikkei.co.jp) — コード/銘柄名/業種/
     ウエイトを持つとされる公開CSV。実エンコーディング・実列構成を確認する。

検証後、本体実装(universe_refresh.py)とともにこのファイル・専用workflowは削除する。
"""

import re
import socket
import sys
import urllib.error
import urllib.request

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

socket.setdefaulttimeout(35)

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")


def fetch(url, timeout=25):
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.status, dict(r.getheaders()), r.read()


def judge_magic(body):
    magic = body[:8]
    if magic[:4] == b'PK\x03\x04':
        return "ZIP系コンテナ → xlsx(OOXML)の可能性 → openpyxlで読める可能性"
    if magic[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return "OLE2/BIFF形式 → 旧xls形式 → xlrdが必要(openpyxl非対応)"
    if body[:1] in (b'\xef', b'\x82', b',') or b',' in body[:200]:
        return "CSVらしきテキスト"
    return f"不明 (先頭8バイトhex={magic.hex()})"


def try_open_excel(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        print(f"  openpyxlで読込成功: sheets={wb.sheetnames}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 3:
                break
            print("   ", row[:12])
        return True
    except Exception as e:
        print(f"  openpyxl失敗: {type(e).__name__}: {e}")
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        print(f"  xlrdで読込成功: rows={ws.nrows} cols={ws.ncols}")
        for i in range(min(4, ws.nrows)):
            print("   ", ws.row_values(i)[:12])
        return True
    except Exception as e:
        print(f"  xlrd失敗: {type(e).__name__}: {e}")
    return False


def probe_jpx_listed_all():
    print("\n=== [1] JPX 東証上場銘柄一覧 (data_j.xls 相当・規模区分でTOPIX500判定できるか) ===")
    idx_urls = [
        "https://www.jpx.co.jp/markets/statistics-equities/misc/01.html",
    ]
    found_links = []
    for idx_url in idx_urls:
        try:
            status, headers, body = fetch(idx_url)
            html = body.decode("utf-8", errors="replace")
            print(f"{idx_url} status={status} bytes={len(body)}")
            links = re.findall(r'href="([^"]*data_j\.xlsx?[^"]*)"', html, re.IGNORECASE)
            links += re.findall(r'href="([^"]+/[^"/"]*上場銘柄一覧[^"]*\.xlsx?)"', html, re.IGNORECASE)
            links = sorted(set(links))
            print(f"  data_j関連リンク数: {len(links)}")
            for l in links[:10]:
                print("   ", l)
            found_links.extend(links)
            if not links:
                # ページ内の.xls/.xlsxリンクを総ざらい
                allx = sorted(set(re.findall(r'href="([^"]+\.xlsx?)"', html, re.IGNORECASE)))
                print(f"  (フォールバック) ページ内の全xls/xlsxリンク数: {len(allx)}")
                for l in allx[:15]:
                    print("   ", l)
                found_links.extend(allx)
        except Exception as e:
            print(f"{idx_url} 失敗: {type(e).__name__}: {e}")

    if not found_links:
        print("data_j.xls相当のリンクが見つからなかった。")
        return

    target = found_links[0]
    if target.startswith("/"):
        target = "https://www.jpx.co.jp" + target
    elif not target.startswith("http"):
        target = "https://www.jpx.co.jp/markets/statistics-equities/misc/" + target
    print(f"\nダウンロード試行: {target}")
    try:
        status2, headers2, body2 = fetch(target)
        print(f"status={status2} bytes={len(body2)} content-type={headers2.get('Content-Type')}")
        print(f"形式判定: {judge_magic(body2)}")
        with open("/tmp/data_j_probe.bin", "wb") as f:
            f.write(body2)
        ok = try_open_excel("/tmp/data_j_probe.bin")
        if ok:
            # 規模区分列・33業種区分列の値の分布を見る (xlrd経路。openpyxlはこの形式を開けないため)
            print("\n  --- 規模区分/33業種区分の値サンプル (xlrd経路) ---")
            try:
                import xlrd
                from collections import Counter
                wb = xlrd.open_workbook("/tmp/data_j_probe.bin")
                ws = wb.sheet_by_index(0)
                header = ws.row_values(0)
                print("   header:", header)
                size_col = next((i for i, h in enumerate(header) if "規模" in str(h) and "コード" not in str(h)), None)
                sector_col = next((i for i, h in enumerate(header) if h == "33業種区分"), None)
                market_col = next((i for i, h in enumerate(header) if "市場" in str(h)), None)
                print(f"   規模区分列={size_col}, 33業種区分列={sector_col}, 市場区分列={market_col}")
                size_cnt, market_cnt = Counter(), Counter()
                for i in range(1, ws.nrows):
                    row = ws.row_values(i)
                    if size_col is not None and len(row) > size_col:
                        size_cnt[row[size_col]] += 1
                    if market_col is not None and len(row) > market_col:
                        market_cnt[row[market_col]] += 1
                print(f"   規模区分 値分布(全{ws.nrows - 1}行):", dict(size_cnt))
                print(f"   市場・商品区分 値分布:", dict(market_cnt))
                if size_col is not None:
                    topix500_n = sum(v for k, v in size_cnt.items()
                                      if str(k) in ("TOPIX Core30", "TOPIX Large70", "TOPIX Mid400"))
                    print(f"   仮説検証: Core30+Large70+Mid400 の合計行数 = {topix500_n} (500に近いか？)")
                # 33業種区分のサンプル(ユニーク値)
                if sector_col is not None:
                    sectors = sorted({ws.row_values(i)[sector_col] for i in range(1, ws.nrows)
                                       if len(ws.row_values(i)) > sector_col})
                    print(f"   33業種区分ユニーク値数={len(sectors)}: {sectors}")
            except Exception as e:
                print(f"   列探索失敗(xlrd経路): {type(e).__name__}: {e}")
    except Exception as e:
        print(f"ダウンロード失敗: {type(e).__name__}: {e}")


def probe_nikkei225():
    print("\n=== [2] 日経225 公式構成銘柄ウエイトCSV ===")
    candidates = [
        "https://indexes.nikkei.co.jp/nkave/archives/file/nikkei_stock_average_weight_jp.csv",
        "https://indexes.nikkei.co.jp/nkave/index?type=download",
        "https://indexes.nikkei.co.jp/nkave/index/component?idx=nk225",
    ]
    for url in candidates:
        try:
            status, headers, body = fetch(url)
            ctype = headers.get("Content-Type", "")
            print(f"\n{url}\n  status={status} bytes={len(body)} content-type={ctype}")
            if "csv" in ctype.lower() or url.endswith(".csv"):
                for enc in ("cp932", "utf-8", "shift_jis"):
                    try:
                        text = body.decode(enc)
                        lines = [l for l in text.splitlines() if l.strip()]
                        print(f"  デコード成功 enc={enc}, 行数(空行除く)={len(lines)}, データ行数={len(lines)-1}")
                        for l in lines[:5]:
                            print("   ", l)
                        print("   ...")
                        for l in lines[-3:]:
                            print("   ", l)
                        break
                    except UnicodeDecodeError:
                        continue
            elif "html" in ctype.lower():
                html = body.decode("utf-8", errors="replace")
                print("  HTML冒頭500字:", re.sub(r"\s+", " ", html[:500]))
                # ページ内にCSVダウンロードリンクがあるか
                links = sorted(set(re.findall(r'href="([^"]*\.csv)"', html, re.IGNORECASE)))
                print(f"  ページ内csvリンク数: {len(links)}")
                for l in links[:10]:
                    print("   ", l)
            else:
                print(f"  先頭200バイト: {body[:200]!r}")
        except urllib.error.HTTPError as e:
            print(f"\n{url}\n  HTTPError: {e.code} {e.reason}")
        except Exception as e:
            print(f"\n{url}\n  失敗: {type(e).__name__}: {e}")


if __name__ == "__main__":
    probe_jpx_listed_all()
    probe_nikkei225()
    print("\n=== プローブ完了 ===")
